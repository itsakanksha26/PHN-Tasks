import openpyxl

workbook = openpyxl.load_workbook("E:\Downloads\FruitsTotal.xlsx")
sheet = workbook.active

fruit_total = {}

# Iterate over each row in the table
for row in sheet.iter_rows(min_row=2, values_only=True):
    fruit = row[0]
    quantity = row[1]

    # Add the quantity to the total for this fruit
    if fruit in fruit_total:
        fruit_total[fruit] += quantity
    else:
        fruit_total[fruit] = quantity

# Create a new Excel spreadsheet
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Write the header row
new_sheet['A1'] = 'Fruit'
new_sheet['B1'] = 'Total Quantity'

# Write the fruit totals to the new spreadsheet
row_index = 2
for fruit, total_quantity in fruit_total.items():
    new_sheet.cell(row=row_index, column=1, value=fruit)
    new_sheet.cell(row=row_index, column=2, value=total_quantity)
    row_index += 1

# Save the new spreadsheet
new_workbook.save("E:\Downloads\FruitsTotal.xlsx")
